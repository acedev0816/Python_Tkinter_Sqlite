#!/usr/bin/env python3

import sys
import tkinter as tk
import sqlite3
import xlrd
import datetime
#from table  import UserTable
import openpyxl ##2.4.0
from openpyxl import Workbook
from ast import literal_eval
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from xlrd.sheet import ctype_text
from tkinter import messagebox
from tkinter import ttk
from tkinter import filedialog
from tkinter.simpledialog import askstring
from tksheet import Sheet


LARGE_FONT = ('Verdana', 18)
SQL_CON = sqlite3.connect('FileParser_DB.db') ## Database file (ADD DATABAE FILE LOCATION HERE)
cursorObj = SQL_CON.cursor()
parent_form={}
sel_id = -1
class Parser(tk.Tk):

    def __init__(self, *args, **kwargs):

        tk.Tk.__init__(self, *args, **kwargs)

        tk.Tk.wm_title(self, "PORTAL")
        self.geometry('500x400')
        self.resizable(False,False)


        container = tk.Frame(self)

        container.pack(side='top', fill='both', expand = True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {} ##contains the dictionary of the frames

        adminstartframe = AdminStartPage(container, self) ## initial frame
        self.frames[AdminStartPage] = adminstartframe

        userstartframe = UserStartPage(container, self)
        self.frames[UserStartPage] = userstartframe

        registerstartframe = RegisterPage(container, self)
        self.frames[RegisterPage] = registerstartframe


        loginframe = LoginPage(container, self)
        self.frames[LoginPage] = loginframe

        userofferframe = UserTable(container,self)
        self.frames[UserTable] = userofferframe
        
        parent_form[Parser] = self
        self.show_frame(LoginPage)

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.grid(row=0, column=0, sticky='nsew')
        if( cont == UserTable):
            frame.showDb()
        frame.tkraise()

#--------------------------------------------------------------------------------------------------------------
class LoginPage(tk.Frame):
    def __init__(self, parent, controller):

        tk.Frame.__init__(self, parent)

        self.parent = parent
        style = ttk.Style()
        style.configure("TButton", padding=6, relief="flat", foreground="#ffa500", background="#ccc")

        label = tk.Label(self, text='LOG IN', font=LARGE_FONT, fg='#ffa500', bg='#808080')
        userLabel = tk.Label(self ,text='User Name',font='50')
        passLabel = tk.Label(self,text='Password',font='50' )

        separator1 = ttk.Separator(self)
        separator2 = ttk.Separator(self)
        separator3 = ttk.Separator(self)
        
        self.usernameStr = tk.StringVar(self)
        self.passwordStr = tk.StringVar(self)

        username = tk.Entry(self, font=LARGE_FONT ,textvariable=self.usernameStr)        
        password = tk.Entry(self, font=LARGE_FONT,textvariable=self.passwordStr)


        loginBtn = ttk.Button(self, text = 'Login', style='TButton', command = self.loginclick)
        registerBtn = ttk.Button(self, text = 'Register', style='TButton', command = self.registerclick)
        

        label.pack(ipadx=320,ipady=10)
        separator1.pack(pady=10)

        userLabel.pack(pady=5,padx=20)
        username.pack(pady=5,padx=20)
        separator2.pack(pady=10)

        passLabel.pack(pady=5,padx=20)
        password.pack(pady=5,padx=20)
        separator3.pack(pady=10)

        loginBtn.pack(ipadx=40,pady=5)
        registerBtn.pack(ipadx=40,pady=5)
    def loginclick(self):
        uName = self.usernameStr.get()
        pwd = self.passwordStr.get()
        sql_query = "SELECT role FROM User_tb WHERE user_name='"+uName+"' AND password='"+pwd+"'"
        print(sql_query)
        cursorObj.execute(sql_query)
        SQL_CON.commit()
        role = cursorObj.fetchall()
        if len(role) > 0:
            role = int(role[0][0])
            if role == 1:
                parent_form[Parser].show_frame(AdminStartPage)
            else:
                parent_form[Parser].show_frame(UserStartPage)
    def registerclick(self):
        parent_form[Parser].show_frame(RegisterPage)


        
        
        

#---------------------------------------------------------------------------------------------------------------

class RegisterPage(tk.Frame):
    
    def __init__(self, parent, controller):
        cursorObj.execute('CREATE TABLE IF NOT EXISTS User_tb(user_id integer PRIMARY KEY,user_name text NOT NULL, password text NOT NULL,role integer)')
        SQL_CON.commit()
        

        tk.Frame.__init__(self, parent)

        style = ttk.Style()
        style.configure("TButton", padding=6, relief="flat", foreground="#ffa500", background="#ccc")

        label = tk.Label(self, text='REGISTER', font=LARGE_FONT, fg='#ffa500', bg='#808080')
        userLabel = tk.Label(self ,text='User Name',font='50')
        passLabel = tk.Label(self,text='Password',font='50' )
        confirmLabel = tk.Label(self,text='Confirm',font='50' )

        separator1 = ttk.Separator(self)
        separator2 = ttk.Separator(self)
        separator3 = ttk.Separator(self)
        separator4 = ttk.Separator(self)

        self.usernameStr = tk.StringVar(self)
        self.passwordStr = tk.StringVar(self)
        self.confirmStr = tk.StringVar(self)

        username = tk.Entry(self, font='100' ,textvariable=self.usernameStr ) 
        password = tk.Entry(self, font='55', textvariable=self.passwordStr )
        confirm = tk.Entry(self, font='55', textvariable=self.confirmStr )


        createBtn = ttk.Button(self, text = 'Create', style='TButton',command = self.createClick)
        cancelBtn = ttk.Button(self, text = 'Back', style='TButton',command = self.cancelClick)
        

        label.pack(ipadx=320,ipady=10)
        separator1.pack(pady=3)

        userLabel.pack(pady=5,padx=20)
        username.pack(ipadx=30,ipady=5,pady=3,padx=20)
        separator2.pack(pady=4)

        passLabel.pack(pady=5,padx=20)
        password.pack(ipadx=30,ipady=5,pady=3,padx=20)
        separator3.pack(pady=3)

        confirmLabel.pack(pady=5,padx=20)
        confirm.pack(ipadx=30,ipady=5,pady=3,padx=20)
        separator4.pack(pady=3)

        createBtn.pack(ipadx=50,pady=5)
        cancelBtn.pack(ipadx=50,pady=5)
    def createClick(self):
        uName = self.usernameStr.get()
        pwd = self.passwordStr.get()
        cpwd = self.confirmStr.get()
        if uName == "":
            MsgBox = tk.messagebox.showinfo("Empty username")
            return
        if pwd == "":
            MsgBox = tk.messagebox.showinfo("Empty password")
            return
        if cpwd == "":
            MsgBox = tk.messagebox.showinfo("Empty confirm")
            return
        if pwd != cpwd:
            MsgBox = tk.messagebox.showinfo("password error")
            return
        cursorObj.execute('SELECT MAX(user_id)FROM User_tb')
        SQL_CON.commit()
        mylist = cursorObj.fetchall()
        if len(mylist) > 0:
            id =int(mylist[0][0]) + 1
            sql_query = "INSERT INTO User_tb VALUES (" + str(id) + ",'" + uName+"','"+pwd +"',0)"
        else:
            cursorObj.execute("INSERT INTO User_tb VALUES (0,'admin','admin',1)")
            SQL_CON.commit()    
            id = 1
            sql_query = "INSERT INTO User_tb VALUES (" + str(id) + ",'" + uName+"','"+pwd +"',0)"            
            
        cursorObj.execute(sql_query)
        SQL_CON.commit()
        parent_form[Parser].show_frame(LoginPage)
    def cancelClick(self):
        parent_form[Parser].show_frame(LoginPage)

#---------------------------------------------------------------------------------------------------------------

class AdminStartPage(tk.Frame):

    def __init__(self, parent, controller):

        cursorObj.execute('CREATE TABLE IF NOT EXISTS project_importance(RAW_Import_date,DocNo PRIMARY KEY,Title,Abstract,Filed_1,Published_1,Opposition_1,Date_of_grant_1,Applicant,Inventor,Legal_Status,PDF_Link,FAMID,Fam,Filed_2,Published_2,Opposition_2,Date_of_grant_2,Highest_legal_status_and_Validity,Topic,Tech_Evaluator,Brief_tech_descr,Eval_Status,Relevance,Relevancy_reason_comments,Work_around_potential)')
        SQL_CON.commit()

        tk.Frame.__init__(self, parent)

        style = ttk.Style()
        style.configure("TButton", padding=6, relief="flat", foreground="#ffa500", background="#ccc")
        label = tk.Label(self, text='APPLICATION', font=LARGE_FONT, fg='#ffa500', bg='#808080')
        separator = ttk.Separator(self)
        loadBtn = ttk.Button(self, text = 'UPDATE CURRNT ANALYSIS WITH RESULTS FROM PAST', style='TButton', command = self.updateFile)
        updateBtn = ttk.Button(self, text = 'Import analysis', style='TButton')
        exportBtn = ttk.Button(self, text = 'Export database')
        offerBtn = ttk.Button(self, text = 'Offer role',command = self.offerClick )
        backBtn = ttk.Button(self, text = 'Back',command = self.backClick )
        label.pack(ipadx=320,ipady=10)
        separator.pack(pady=35)

        loadBtn.pack(pady=5)
        updateBtn.pack(pady=5)
        exportBtn.pack(pady=5)
        offerBtn.pack(ipadx = 10, pady=5)
        backBtn.pack(ipadx = 10, pady=5)
        
    def offerClick(self):
        parent_form[Parser].show_frame(UserTable)
    def backClick(self):
        parent_form[Parser].show_frame(LoginPage)

# --------------------------------------------------------------------------------------------------------------------------------------------------------- BUTTONS

# ----------------------------------------------------------------------------------------------------------------------------------------- Update File
    def updateFile(self): ## Adds old fields of Relevance and RelevancyComments from DB if the projects are found
        try:
            # ----------------------------------------------------------------------------------------------------------------------------------------- User inputs
            self.filename = filedialog.askopenfilename(initialdir = '/', title = 'Select a file', filetype = (('Excel files', '*.xl*'), ('All files', '*.*')))

            if self.filename != '':
                MsgBox = tk.messagebox.askquestion ('Import','Are you sure you want to update "' + self.filename.split('/')[-1:][0] + '"',icon = 'question')
                if MsgBox == 'yes':
                    print('\nREADING FILE...\n')

                    wb = openpyxl.load_workbook(self.filename)
                    ws = wb.active


                    # ----------------------------------------------------------------------------------------------------------------------------------------- Color box
                    for cells in ws.iter_cols(min_row=3, min_col=1, max_col=3, max_row=12):
                        for cell in cells:
                            cell.fill = PatternFill(start_color='00000000', end_color='00000000', fill_type='solid')
                            cell.font = Font(bold=True, color='FFFFFFFF')

                    startingRow=0
                    docColumn=''
                    i=1
                    # ----------------------------------------------------------------------------------------------------------------------------------------- Find headings
                    for cell in ws['B']:
                        if cell.value == 'DocNo' or cell.value == 'document' or cell.value == 'Document':
                            startingRow = i
                            docColumn = 'B'
                        i+=1
                    j=1
                    for cell in ws['A']:
                        if cell.value == 'document' or cell.value == 'DocNo' or cell.value == 'Document':
                            startingRow = j
                            docColumn = 'A'
                        j+=1

                    if startingRow == 0:
                        print("Couldn't find DocNo/document, it should be in column 1 or column 2")


                    headings = []
                    for heading in ws.iter_cols(min_row=startingRow, min_col=0, max_col=30, max_row=startingRow):
                      for cell in heading:
                          if cell.value:
                            headings.append(cell.value)

                    # ----------------------------------------------------------------------------------------------------------------------------------------- Parse file again
                    if headings[len(headings)-1] == 'Old Relevancy reason comments':
                        print('Updating file again...')
                        relLetter = openpyxl.utils.cell.get_column_letter(len(headings)-1)
                        rel2Letter = openpyxl.utils.cell.get_column_letter(len(headings))
                    else:
                        relLetter = openpyxl.utils.cell.get_column_letter(len(headings)+1)
                        rel2Letter = openpyxl.utils.cell.get_column_letter(len(headings)+2)

                    # print('%s rl - %s sr' %(relLetter,startingRow))
                    ws[relLetter + str(startingRow)] = 'Old Relevance'
                    ws[rel2Letter + str(startingRow)] = 'Old Relevancy reason comments'


                    i = startingRow+1
                    # ----------------------------------------------------------------------------------------------------------------------------------------- Iterate rows
                    while ws[docColumn + str(i)].value != None:

                        DocNumber = ws[docColumn + str(i)].value.replace(' ','')

                        cursorObj.execute('SELECT Relevance,Relevancy_reason_comments FROM project_importance WHERE "DocNo" = "' + DocNumber + '"')
                        oldValues = cursorObj.fetchall()
                        if len(oldValues) > 0:
                            try:
                                oldValues = oldValues[0]
                                if oldValues[0] != 'NULL':
                                    ws[relLetter + str(i)] = str(oldValues[0])
                                    print("%s Relevance has been UPDATED"%(DocNumber))
                                if oldValues[1] != 'NULL':
                                    ws[rel2Letter + str(i)] = str(oldValues[1])
                                    print("%s Relevancy_reason_comments has been UPDATED"%(DocNumber))
                            except Exception as e:
                                print('Error reading the file %s - %s' %(e, sys.exc_info()[2].tb_lineno))
                                pass
                        else:
                            print('No line with docNo - %s' %(DocNumber.split(' ')[1]))
                        i+=1

                    wb.save(self.filename)

                    print('"%s" has been UPDATED' %(self.filename))

        except Exception as e:
            tb = sys.exc_info()[2]
            print('An error ocurred in line %s - %s' %(e, tb.tb_lineno))
#-----------------------------------------------------------------------------------------------------------------------------
class UserStartPage(tk.Frame):

    def __init__(self, parent, controller):

        cursorObj.execute('CREATE TABLE IF NOT EXISTS project_importance(RAW_Import_date,DocNo PRIMARY KEY,Title,Abstract,Filed_1,Published_1,Opposition_1,Date_of_grant_1,Applicant,Inventor,Legal_Status,PDF_Link,FAMID,Fam,Filed_2,Published_2,Opposition_2,Date_of_grant_2,Highest_legal_status_and_Validity,Topic,Tech_Evaluator,Brief_tech_descr,Eval_Status,Relevance,Relevancy_reason_comments,Work_around_potential)')
        SQL_CON.commit()

        tk.Frame.__init__(self, parent)

        style = ttk.Style()
        style.configure("TButton", padding=6, relief="flat", foreground="#ffa500", background="#ccc")
        label = tk.Label(self, text='APPLICATION', font=LARGE_FONT, fg='#ffa500', bg='#808080')
        separator = ttk.Separator(self)
        exportBtn = ttk.Button(self, text = 'Export database')

        label.pack(ipadx=320,ipady=10)
        separator.pack(pady=35)
        exportBtn.pack(pady=5)

#----------------------------------------------------------------------------------------------------------------------------------------------------------------
class UserTable(tk.Frame):
    
    def __init__(self, parent, controller):
        
        tk.Frame.__init__(self,parent)

        style = ttk.Style()
        style.configure("TButton", padding=6, relief="flat", foreground="#ffa500", background="#ccc")
        
        btnContain = tk.Label(self)

        updateBtn = ttk.Button(btnContain, text = 'Offer', style='TButton',command = self.offerClick)
        backBtn = ttk.Button(btnContain, text = 'Back', style='TButton',command = self.backClick)

        updateBtn.pack(side = tk.LEFT, expand = True, ipadx=40,padx=10)
        backBtn.pack(side = tk.LEFT, expand = True, ipadx=40,padx=10)
        parent_form[UserTable] = self

        self.grid_columnconfigure(0, weight = 1)
        self.grid_rowconfigure(0, weight = 1)
        self.sheet_user = Sheet(self,
                                #data = [[f"Row {r} Column {c}" for c in range(30)] for r in range(200)], #to set sheet data at startup
                                height = 300,
                                width = 400) #For full startup arguments/parameters see DOCUMENTATION.md
        self.sheet_user.enable_bindings(("single_select", #"single_select" or "toggle_select"
                                        # "drag_select",   #enables shift click selection as well
                                        # "column_drag_and_drop",
                                         "row_drag_and_drop",
                                        #"column_select",
                                         "row_select",
                                         #"column_width_resize",
                                         #"double_click_column_resize",
                                         "row_width_resize",
                                         "column_height_resize",
                                         "arrowkeys",
                                         "row_height_resize",
                                         #"double_click_row_resize",
                                         #"right_click_popup_menu",
                                         #"rc_insert_column",
                                         #"rc_delete_column",
                                         #"rc_insert_row",
                                         #"rc_delete_row",
                                         #"copy",
                                         #"cut",
                                         #"paste",
                                         #"delete",
                                         #"undo",
                                         #"edit_cell"
                                         ))
        #self.sheet_user.disable_bindings() #uses the same strings
        self.sheet_user.grid(row = 0, column = 0, sticky = "we")
        
        btnContain.grid(row = 1, column = 0, sticky = "nswe")
        """_________________________ EXAMPLES _________________________ """
        """_____________________________________________________________"""

        # __________ CHANGING THEME __________

#        self.sheet_user.change_theme("dark")

        # __________ HIGHLIGHT / DEHIGHLIGHT CELLS __________
        
        #self.sheet_user.highlight_cells(row = 5, column = 5, bg = "#ed4337", fg = "white")
        #self.sheet_user.highlight_cells(row = 5, column = 1, bg = "#ed4337", fg = "white")
        #self.sheet_user.highlight_cells(row = 5, bg = "#ed4337", fg = "white", canvas = "row_index")
        #self.sheet_user.highlight_cells(column = 0, bg = "#ed4337", fg = "white", canvas = "header")

        # __________ SETTING OR RESETTING TABLE DATA __________

        #.set_sheet_data() function returns the object you use as argument
        #verify checks if your data is a list of lists, raises error if not
           
        # __________ BINDING A FUNCTION TO USER SELECTS CELL __________

        self.sheet_user.extra_bindings([
                                        ("cell_select", self.cell_select),
                                        ("shift_cell_select", self.shift_select_cells),
                                        ("drag_select_cells", self.drag_select_cells),
                                        ("ctrl_a", self.ctrl_a),
                                        ("row_select", self.row_select),
                                        ("shift_row_select", self.shift_select_rows),
                                        ("drag_select_rows", self.drag_select_rows),
                                        ("column_select", self.column_select),
                                        ("shift_column_select", self.shift_select_columns),
                                        ("drag_select_columns", self.drag_select_columns),
                                        ("deselect", self.deselect)
                                        ]
                                       )
        
        # __________ BINDING NEW RIGHT CLICK FUNCTION __________
       
           
        #self.sheet_user.bind("<3>", self.rc)

        # __________ DISPLAY SUBSET OF COLUMNS __________

        #self.sheet_user.display_subset_of_columns(indexes = [3, 7, 9, 0], enable = True)

        # __________ SETTING HEADERS __________

        self.sheet_user.headers((f"Header {a}" for a in range(4))) #any iterable works
        self.sheet_user.headers("Id", 0)
        self.sheet_user.headers("User Name", 1)
        self.sheet_user.headers("Password", 2)
        self.sheet_user.headers("Role", 3)
        #print (self.sheet_user.headers())
        #print (self.sheet_user.headers(index = 2))

        # __________ SETTING ROW INDEX __________

        #self.sheet_user.row_index((f"Row {r}" for r in range(200))) #any iterable works
        #self.sheet_user.row_index("Change index example", 2)
        #print (self.sheet_user.row_index())
        #print (self.sheet_user.row_index(index = 2))

        # __________ INSERTING A ROW __________

        #self.sheet_user.insert_row(row = (f"my new row here {c}" for c in range(100)), idx = 0) # a filled row at the start
        #self.sheet_user.insert_row() # an empty row at the end

        # __________ INSERTING A COLUMN __________

        #self.sheet_user.insert_column(column = (f"my new col here {r}" for r in range(5000)), idx = 0) # a filled column at the start
        #self.sheet_user.insert_column() # an empty column at the end

        # __________ SETTING A COLUMNS DATA __________

        # any iterable works
        #self.sheet_user.set_column_data(0, values = (0 for i in range(220)))

        # __________ SETTING A ROWS DATA __________

        # any iterable works
        #self.sheet_user.set_row_data(0, values = (0 for i in range(35)))



        # __________ SETTING A CELLS DATA __________

        #self.sheet_user.set_cell_data(1, 2, "NEW VALUE")

        # __________ GETTING FULL SHEET DATA __________

        #self.all_data = self.sheet_user.get_sheet_data()

        # __________ GETTING CELL DATA __________

        #print (self.sheet_user.get_cell_data(0, 0))

        # __________ GETTING ROW DATA __________

        #print (self.sheet_user.get_row_data(0)) # only accessible by index

        # __________ GETTING COLUMN DATA __________

        #print (self.sheet_user.get_column_data(0)) # only accessible by index

        # __________ HIDING THE ROW INDEX AND HEADERS __________

        #self.sheet_user.hide("row_index")
        #self.sheet_user.hide("top_left")
        #self.sheet_user.hide("header")
        self.showDb()
   
    def showDb(self):
        cursorObj.execute('SELECT * FROM User_tb')
        SQL_CON.commit()
        mylist = cursorObj.fetchall()
        count = len(mylist)
        self.data = self.sheet_user.set_sheet_data([[f"Row {r} Column {c}" for c in range(3)] for r in range(count)], verify = False)

        i = 0
        for row in mylist:
        #    if i > 0: 
            lst = list(row)
            if(lst[3] == 0):
                lst[3] = 'general'
            else:
                lst[3] = 'admin'
            row = tuple(lst)
            self.sheet_user.set_row_data(i,row)
            i += 1    
        #print(self)
        self.sheet_user.refresh()
    def deselect(self, event):
        print (event, self.sheet_user.get_selected_cells())

    def rc(self, event):
        print (event)
        
    def cell_select(self, response):
        global sel_id
        sel_id = response[1]
        self.sheet_user.select_row(sel_id);
        print (sel_id)

    def shift_select_cells(self, response):
        print (response)

    def drag_select_cells(self, response):
        pass
        #print (response)

    def ctrl_a(self, response):
        print (response)

    def row_select(self, response):
        global sel_id
        sel_id = response[1]
        print (sel_id)

    def shift_select_rows(self, response):
        print (response)

    def drag_select_rows(self, response):
        pass
        #print (response)
        
    def column_select(self, response):
        print (response)

    def shift_select_columns(self, response):
        print (response)

    def drag_select_columns(self, response):
        pass
    def offerClick(btnContain):
        global sel_id
        if sel_id == -1:
            return
        cur_id = parent_form[UserTable].sheet_user.get_cell_data(sel_id, 0)
        
        query = "UPDATE User_tb SET role = 1 WHERE user_id = " + cur_id;

        cursorObj.execute(query)
        SQL_CON.commit()

        parent_form[UserTable].sheet_user.set_cell_data(sel_id, 3, "admin")
        parent_form[UserTable].sheet_user.refresh()

    def backClick(btnContain):
        parent_form[Parser].show_frame(AdminStartPage)
        #print (response) 

def before_closing():
    MsgBox = tk.messagebox.askquestion ('Exit Application','Are you sure you want to exit the application',icon = 'warning')
    if MsgBox == 'yes':
        SQL_CON.close()
        app.destroy()


if __name__ == "__main__":
    app = Parser()
    app.protocol("WM_DELETE_WINDOW", before_closing)
    print('\n')
    print('    #######################################################################')
    print('    #     _     _  _______  ___      _______  _______  __   __  _______   #')
    print('    #    | | _ | ||       ||   |    |       ||       ||  |_|  ||       |  #')
    print('    #    | || || ||    ___||   |    |       ||   _   ||       ||    ___|  #')
    print('    #    |       ||   |___ |   |    |       ||  | |  ||       ||   |___   #')
    print('    #    |       ||    ___||   |___ |      _||  |_|  ||       ||    ___|  #')
    print('    #    |   _   ||   |___ |       ||     |_ |       || ||_|| ||   |___   #')
    print('    #    |__| |__||_______||_______||_______||_______||_|   |_||_______|  #')
    print('    #                                                                     #')
    print('    #######################################################################')
    print('    #')
    print('    #----> Patent database manager\n')
    print('Starting GUI...')
    app.mainloop()
